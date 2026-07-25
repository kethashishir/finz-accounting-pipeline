"""Read-only inspection of untrusted CSV and XLSX source files."""

from __future__ import annotations

import csv
import hashlib
import io
import zipfile
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.models.ingestion import FileType
from app.models.source import (
    SourceFileInspection,
    SourcePreviewRow,
    SourceSheetPreview,
)


class SourceInspectionErrorCode(StrEnum):
    """Stable machine-readable source-inspection errors."""

    EMPTY_FILE = "empty_file"
    FILE_TOO_LARGE = "file_too_large"
    UNSUPPORTED_EXTENSION = "unsupported_extension"
    SIGNATURE_MISMATCH = "signature_mismatch"
    INVALID_CSV = "invalid_csv"
    INVALID_XLSX = "invalid_xlsx"
    UNSAFE_ARCHIVE = "unsafe_archive"
    TOO_MANY_COLUMNS = "too_many_columns"


class SourceInspectionError(ValueError):
    """An uploaded file cannot be inspected safely."""

    def __init__(
        self,
        code: SourceInspectionErrorCode,
        message: str,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


class SourceFileInspector:
    """Inspect source structure without executing or modifying content."""

    CSV_SHEET_NAME = "CSV"
    XLSX_SIGNATURE = b"PK\x03\x04"

    def __init__(
        self,
        *,
        max_file_bytes: int = 10 * 1024 * 1024,
        max_uncompressed_bytes: int = 50 * 1024 * 1024,
        max_zip_entries: int = 1_000,
        max_columns: int = 100,
        preview_row_count: int = 10,
    ) -> None:
        self.max_file_bytes = max_file_bytes
        self.max_uncompressed_bytes = max_uncompressed_bytes
        self.max_zip_entries = max_zip_entries
        self.max_columns = max_columns
        self.preview_row_count = preview_row_count

    def inspect(
        self,
        *,
        file_name: str,
        content: bytes,
    ) -> SourceFileInspection:
        """Return structural metadata and preview rows for an upload."""

        if not content:
            raise SourceInspectionError(
                SourceInspectionErrorCode.EMPTY_FILE,
                "The uploaded file is empty",
            )

        if len(content) > self.max_file_bytes:
            raise SourceInspectionError(
                SourceInspectionErrorCode.FILE_TOO_LARGE,
                f"The uploaded file exceeds {self.max_file_bytes} bytes",
            )

        safe_file_name = self._safe_file_name(file_name)
        file_type = self._detect_file_type(safe_file_name)
        file_sha256 = hashlib.sha256(content).hexdigest()

        if file_type == FileType.CSV:
            return self._inspect_csv(
                original_file_name=file_name,
                safe_file_name=safe_file_name,
                content=content,
                file_sha256=file_sha256,
            )

        return self._inspect_xlsx(
            original_file_name=file_name,
            safe_file_name=safe_file_name,
            content=content,
            file_sha256=file_sha256,
        )

    @staticmethod
    def _safe_file_name(file_name: str) -> str:
        """Remove browser-supplied or platform path components."""

        normalized = file_name.replace("\\", "/")
        safe_name = Path(normalized).name.strip()

        if not safe_name:
            raise SourceInspectionError(
                SourceInspectionErrorCode.UNSUPPORTED_EXTENSION,
                "The uploaded file name is empty",
            )

        return safe_name

    @staticmethod
    def _detect_file_type(file_name: str) -> FileType:
        """Allow only explicitly supported source extensions."""

        suffix = Path(file_name).suffix.lower()

        if suffix == ".csv":
            return FileType.CSV
        if suffix == ".xlsx":
            return FileType.XLSX

        raise SourceInspectionError(
            SourceInspectionErrorCode.UNSUPPORTED_EXTENSION,
            "Only .csv and .xlsx source files are supported",
        )

    def _inspect_csv(
        self,
        *,
        original_file_name: str,
        safe_file_name: str,
        content: bytes,
        file_sha256: str,
    ) -> SourceFileInspection:
        """Inspect CSV encoding, delimiter, shape, and preview rows."""

        if content.startswith(self.XLSX_SIGNATURE):
            raise SourceInspectionError(
                SourceInspectionErrorCode.SIGNATURE_MISMATCH,
                "File content is an XLSX archive but the extension is .csv",
            )

        text, encoding, warnings = self._decode_csv(content)
        delimiter = self._detect_delimiter(text, warnings)

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        preview_rows: list[SourcePreviewRow] = []
        row_count = 0
        column_count = 0

        try:
            for row_number, row in enumerate(reader, start=1):
                row_count = row_number
                column_count = max(column_count, len(row))

                if len(row) > self.max_columns:
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.TOO_MANY_COLUMNS,
                        f"CSV row {row_number} exceeds "
                        f"{self.max_columns} columns",
                    )

                if row_number <= self.preview_row_count:
                    preview_rows.append(
                        SourcePreviewRow(
                            row_number=row_number,
                            values=row,
                        )
                    )
        except csv.Error as exc:
            raise SourceInspectionError(
                SourceInspectionErrorCode.INVALID_CSV,
                f"CSV parsing failed: {exc}",
            ) from exc

        if row_count == 0:
            raise SourceInspectionError(
                SourceInspectionErrorCode.INVALID_CSV,
                "The CSV file contains no rows",
            )

        sheet = SourceSheetPreview(
            name=self.CSV_SHEET_NAME,
            row_count=row_count,
            column_count=column_count,
            preview_rows=preview_rows,
        )

        return SourceFileInspection(
            original_file_name=original_file_name,
            safe_file_name=safe_file_name,
            file_type=FileType.CSV,
            file_sha256=file_sha256,
            size_bytes=len(content),
            encoding=encoding,
            delimiter=delimiter,
            sheets=[sheet],
            warnings=warnings,
        )

    @staticmethod
    def _decode_csv(content: bytes) -> tuple[str, str, list[str]]:
        """Decode common bank-export encodings without dropping bytes."""

        try:
            return content.decode("utf-8-sig"), "utf-8-sig", []
        except UnicodeDecodeError:
            text = content.decode("cp1252")
            return (
                text,
                "cp1252",
                [
                    "CSV is not UTF-8; decoded as Windows-1252. "
                    "Review non-ASCII text before approval."
                ],
            )

    @staticmethod
    def _detect_delimiter(text: str, warnings: list[str]) -> str:
        """Detect a common delimiter, falling back visibly to comma."""

        sample = text[:8_192]

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=",;\t|",
            )
            return dialect.delimiter
        except csv.Error:
            warnings.append(
                "CSV delimiter could not be detected reliably; "
                "comma was selected as the fallback."
            )
            return ","

    def _inspect_xlsx(
        self,
        *,
        original_file_name: str,
        safe_file_name: str,
        content: bytes,
        file_sha256: str,
    ) -> SourceFileInspection:
        """Inspect an XLSX archive and its worksheets read-only."""

        if not content.startswith(self.XLSX_SIGNATURE):
            raise SourceInspectionError(
                SourceInspectionErrorCode.SIGNATURE_MISMATCH,
                "File extension is .xlsx but content is not an XLSX archive",
            )

        warnings = self._validate_xlsx_archive(content)

        try:
            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (
            InvalidFileException,
            KeyError,
            OSError,
            ValueError,
            zipfile.BadZipFile,
        ) as exc:
            raise SourceInspectionError(
                SourceInspectionErrorCode.INVALID_XLSX,
                "The XLSX workbook could not be parsed safely",
            ) from exc

        sheets: list[SourceSheetPreview] = []

        try:
            for worksheet in workbook.worksheets:
                declared_columns = worksheet.max_column or 0
                if declared_columns > self.max_columns:
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.TOO_MANY_COLUMNS,
                        f"Worksheet '{worksheet.title}' exceeds "
                        f"{self.max_columns} columns",
                    )

                preview_rows: list[SourcePreviewRow] = []
                row_count = 0
                column_count = 0
                formula_count = 0

                for row_number, cells in enumerate(
                    worksheet.iter_rows(),
                    start=1,
                ):
                    row_count = row_number
                    values: list[Any] = []

                    for cell in cells:
                        values.append(cell.value)
                        if cell.data_type == "f":
                            formula_count += 1

                    column_count = max(column_count, len(values))

                    if row_number <= self.preview_row_count:
                        preview_rows.append(
                            SourcePreviewRow(
                                row_number=row_number,
                                values=values,
                            )
                        )

                if formula_count:
                    warnings.append(
                        f"Worksheet '{worksheet.title}' contains "
                        f"{formula_count} formula cell(s). Formulas will not "
                        "be executed and require manual review."
                    )

                sheets.append(
                    SourceSheetPreview(
                        name=worksheet.title,
                        row_count=row_count,
                        column_count=column_count,
                        preview_rows=preview_rows,
                    )
                )
        finally:
            workbook.close()

        if not sheets:
            raise SourceInspectionError(
                SourceInspectionErrorCode.INVALID_XLSX,
                "The XLSX workbook contains no worksheets",
            )

        return SourceFileInspection(
            original_file_name=original_file_name,
            safe_file_name=safe_file_name,
            file_type=FileType.XLSX,
            file_sha256=file_sha256,
            size_bytes=len(content),
            sheets=sheets,
            warnings=warnings,
        )

    def _validate_xlsx_archive(self, content: bytes) -> list[str]:
        """Reject ZIP bombs, macros, and structurally invalid workbooks."""

        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                entries = archive.infolist()

                if len(entries) > self.max_zip_entries:
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.UNSAFE_ARCHIVE,
                        "The XLSX archive contains too many entries",
                    )

                uncompressed_size = sum(entry.file_size for entry in entries)
                if uncompressed_size > self.max_uncompressed_bytes:
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.UNSAFE_ARCHIVE,
                        "The XLSX archive expands beyond the safe size limit",
                    )

                names = {entry.filename for entry in entries}
                lowercase_names = {name.lower() for name in names}

                required_names = {
                    "[Content_Types].xml",
                    "xl/workbook.xml",
                }
                if not required_names.issubset(names):
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.INVALID_XLSX,
                        "The XLSX archive is missing required workbook files",
                    )

                if any(
                    name.endswith("vbaproject.bin")
                    for name in lowercase_names
                ):
                    raise SourceInspectionError(
                        SourceInspectionErrorCode.UNSAFE_ARCHIVE,
                        "Macro-enabled workbooks are not accepted",
                    )

                warnings: list[str] = []
                if any(
                    name.startswith("xl/externallinks/")
                    for name in lowercase_names
                ):
                    warnings.append(
                        "Workbook contains external workbook links. "
                        "Linked values require manual review."
                    )

                return warnings
        except zipfile.BadZipFile as exc:
            raise SourceInspectionError(
                SourceInspectionErrorCode.INVALID_XLSX,
                "The XLSX content is not a valid ZIP archive",
            ) from exc
