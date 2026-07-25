"""Configurable parsing and normalization of bank-export records."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from typing import Any
from uuid import UUID

from openpyxl import load_workbook

from app.models.ingestion import (
    CENT,
    FileType,
    IngestionConfig,
    IssueSeverity,
    NormalizedTransaction,
    RawRecord,
    RecordStatus,
    TransactionDirection,
    ValidationIssue,
)
from app.models.source import SourceFileInspection
from app.services.ingestion.inspector import SourceFileInspector

MONEY_PATTERN = re.compile(r"^[+-]?(?:\d+(?:,\d{3})*(?:\.\d{1,2})?|\d*\.\d{1,2})$")


class SourceNormalizationErrorCode(StrEnum):
    """Stable errors that prevent the entire file from being normalized."""

    FILE_TYPE_MISMATCH = "file_type_mismatch"
    WORKSHEET_REQUIRED = "worksheet_required"
    WORKSHEET_NOT_FOUND = "worksheet_not_found"
    HEADER_ROW_NOT_FOUND = "header_row_not_found"
    INVALID_HEADER = "invalid_header"
    DUPLICATE_HEADER = "duplicate_header"
    MISSING_MAPPED_COLUMN = "missing_mapped_column"
    NO_DATA_ROWS = "no_data_rows"


class SourceNormalizationError(ValueError):
    """A file-level structure or mapping problem."""

    def __init__(
        self,
        code: SourceNormalizationErrorCode,
        message: str,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True, slots=True)
class NormalizationResult:
    """Result of normalizing one inspected upload."""

    inspection: SourceFileInspection
    raw_records: tuple[RawRecord, ...]
    transactions: tuple[NormalizedTransaction, ...]


class SourceFileNormalizer:
    """Create raw and normalized records without silently dropping rows."""

    def __init__(
        self,
        inspector: SourceFileInspector | None = None,
    ) -> None:
        self.inspector = inspector or SourceFileInspector()

    def normalize(
        self,
        *,
        upload_id: UUID,
        file_name: str,
        content: bytes,
        config: IngestionConfig,
    ) -> NormalizationResult:
        """Inspect, parse, and normalize every physical data row."""

        inspection = self.inspector.inspect(
            file_name=file_name,
            content=content,
        )

        if inspection.file_type != config.file_type:
            raise SourceNormalizationError(
                SourceNormalizationErrorCode.FILE_TYPE_MISMATCH,
                "The ingestion configuration does not match the uploaded file type",
            )

        source_sheet, rows = self._load_rows(
            content=content,
            config=config,
            inspection=inspection,
        )

        if config.header_row > len(rows):
            raise SourceNormalizationError(
                SourceNormalizationErrorCode.HEADER_ROW_NOT_FOUND,
                f"Header row {config.header_row} does not exist in the source",
            )

        headers = self._parse_headers(rows[config.header_row - 1][1])
        self._validate_mapped_columns(headers, config)

        data_rows = rows[config.header_row :]
        if not data_rows:
            raise SourceNormalizationError(
                SourceNormalizationErrorCode.NO_DATA_ROWS,
                "The selected header row has no data rows after it",
            )

        raw_records: list[RawRecord] = []
        transactions: list[NormalizedTransaction] = []

        for row_number, values in data_rows:
            raw_values, structural_issues = self._build_raw_values(
                headers=headers,
                values=values,
            )

            raw_record = RawRecord(
                upload_id=upload_id,
                source_file_name=inspection.safe_file_name,
                source_sheet=source_sheet,
                source_row_number=row_number,
                raw_values=raw_values,
                raw_hash=self._raw_hash(raw_values),
            )

            transaction = self._normalize_row(
                upload_id=upload_id,
                raw_record=raw_record,
                config=config,
                structural_issues=structural_issues,
            )

            raw_records.append(raw_record)
            transactions.append(transaction)

        return NormalizationResult(
            inspection=inspection,
            raw_records=tuple(raw_records),
            transactions=tuple(transactions),
        )

    @staticmethod
    def _load_csv_rows(
        *,
        content: bytes,
        inspection: SourceFileInspection,
    ) -> list[tuple[int, list[Any]]]:
        encoding = inspection.encoding or "utf-8-sig"
        delimiter = inspection.delimiter or ","
        text = content.decode(encoding)

        return [
            (row_number, list(row))
            for row_number, row in enumerate(
                csv.reader(io.StringIO(text), delimiter=delimiter),
                start=1,
            )
        ]

    def _load_rows(
        self,
        *,
        content: bytes,
        config: IngestionConfig,
        inspection: SourceFileInspection,
    ) -> tuple[str | None, list[tuple[int, list[Any]]]]:
        if inspection.file_type == FileType.CSV:
            return None, self._load_csv_rows(
                content=content,
                inspection=inspection,
            )

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )

        try:
            if config.sheet_name is None:
                if len(workbook.sheetnames) != 1:
                    raise SourceNormalizationError(
                        SourceNormalizationErrorCode.WORKSHEET_REQUIRED,
                        "Select a worksheet when an XLSX file has multiple sheets",
                    )
                sheet_name = workbook.sheetnames[0]
            else:
                sheet_name = config.sheet_name

            if sheet_name not in workbook.sheetnames:
                raise SourceNormalizationError(
                    SourceNormalizationErrorCode.WORKSHEET_NOT_FOUND,
                    f"Worksheet '{sheet_name}' does not exist",
                )

            worksheet = workbook[sheet_name]
            rows = [
                (row_number, list(values))
                for row_number, values in enumerate(
                    worksheet.iter_rows(values_only=True),
                    start=1,
                )
            ]
        finally:
            workbook.close()

        return sheet_name, rows

    @staticmethod
    def _parse_headers(values: list[Any]) -> list[str]:
        headers: list[str] = []
        seen: set[str] = set()

        for column_number, value in enumerate(values, start=1):
            header = "" if value is None else str(value).strip()

            if not header:
                raise SourceNormalizationError(
                    SourceNormalizationErrorCode.INVALID_HEADER,
                    f"Header column {column_number} is blank",
                )

            comparison_key = header.casefold()
            if comparison_key in seen:
                raise SourceNormalizationError(
                    SourceNormalizationErrorCode.DUPLICATE_HEADER,
                    f"Header '{header}' appears more than once",
                )

            seen.add(comparison_key)
            headers.append(header)

        if not headers:
            raise SourceNormalizationError(
                SourceNormalizationErrorCode.INVALID_HEADER,
                "The selected header row is empty",
            )

        return headers

    @staticmethod
    def _validate_mapped_columns(
        headers: list[str],
        config: IngestionConfig,
    ) -> None:
        available = set(headers)
        mapped_columns = config.column_mapping.model_dump(exclude_none=True).values()

        for mapped_column in mapped_columns:
            if mapped_column not in available:
                raise SourceNormalizationError(
                    SourceNormalizationErrorCode.MISSING_MAPPED_COLUMN,
                    f"Mapped source column '{mapped_column}' does not exist",
                )

    @staticmethod
    def _build_raw_values(
        *,
        headers: list[str],
        values: list[Any],
    ) -> tuple[dict[str, Any], list[ValidationIssue]]:
        raw_values = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        }
        issues: list[ValidationIssue] = []

        extra_values = values[len(headers) :]
        if extra_values:
            for offset, value in enumerate(extra_values, start=1):
                raw_values[f"__extra_column_{offset}"] = value

            issues.append(
                ValidationIssue(
                    code="extra_columns",
                    field="_record",
                    message=("The row contains values beyond the configured header columns"),
                    raw_value=extra_values,
                )
            )

        return raw_values, issues

    def _normalize_row(
        self,
        *,
        upload_id: UUID,
        raw_record: RawRecord,
        config: IngestionConfig,
        structural_issues: list[ValidationIssue],
    ) -> NormalizedTransaction:
        mapping = config.column_mapping
        values = raw_record.raw_values
        issues = list(structural_issues)

        source_transaction_id = self._optional_text(
            values.get(mapping.source_transaction_id) if mapping.source_transaction_id else None
        )

        transaction_date = self._date_field(
            raw_value=values.get(mapping.transaction_date),
            field="transaction_date",
            date_format=config.date_format,
            required=True,
            issues=issues,
        )

        posted_date = self._date_field(
            raw_value=(values.get(mapping.posted_date) if mapping.posted_date else None),
            field="posted_date",
            date_format=config.date_format,
            required=False,
            issues=issues,
        )

        description_value = values.get(mapping.description)
        description_original = None if self._is_blank(description_value) else str(description_value)
        description_normalized = self._normalize_description(description_original)

        if description_normalized is None:
            issues.append(
                ValidationIssue(
                    code="missing_description",
                    field="description",
                    message="Transaction description is required",
                    raw_value=description_value,
                )
            )

        amount = self._amount_field(
            values=values,
            config=config,
            issues=issues,
        )

        currency = self._currency_field(
            values=values,
            config=config,
            issues=issues,
        )

        bank_account = self._bank_account_field(
            values=values,
            config=config,
            issues=issues,
        )

        if (
            transaction_date is not None
            and posted_date is not None
            and posted_date < transaction_date
        ):
            issues.append(
                ValidationIssue(
                    code="posted_before_transaction",
                    field="posted_date",
                    message="Posted date is earlier than transaction date",
                    severity=IssueSeverity.WARNING,
                    raw_value=values.get(mapping.posted_date),
                )
            )

        direction: TransactionDirection | None = None
        if amount is not None and amount != Decimal("0.00"):
            direction = (
                TransactionDirection.INFLOW
                if amount > Decimal("0.00")
                else TransactionDirection.OUTFLOW
            )

        fingerprint = self._fingerprint(
            transaction_date=transaction_date,
            amount=amount,
            currency=currency,
            bank_account=bank_account,
            description=description_normalized,
        )

        has_errors = any(issue.severity == IssueSeverity.ERROR for issue in issues)
        status = RecordStatus.INVALID if has_errors else RecordStatus.VALID

        return NormalizedTransaction(
            upload_id=upload_id,
            raw_record_id=raw_record.id,
            source_transaction_id=source_transaction_id,
            transaction_date=transaction_date,
            posted_date=posted_date,
            description_original=description_original,
            description_normalized=description_normalized,
            amount=amount,
            currency=currency,
            bank_account=bank_account,
            direction=direction,
            fingerprint=fingerprint,
            status=status,
            validation_issues=issues,
        )

    def _amount_field(
        self,
        *,
        values: dict[str, Any],
        config: IngestionConfig,
        issues: list[ValidationIssue],
    ) -> Decimal | None:
        mapping = config.column_mapping

        if mapping.amount is not None:
            amount = self._parse_money_issue(
                raw_value=values.get(mapping.amount),
                field="amount",
                issues=issues,
            )
        else:
            debit_value = values.get(mapping.debit_amount) if mapping.debit_amount else None
            credit_value = values.get(mapping.credit_amount) if mapping.credit_amount else None

            has_debit = not self._is_blank(debit_value)
            has_credit = not self._is_blank(credit_value)

            if has_debit and has_credit:
                issues.append(
                    ValidationIssue(
                        code="ambiguous_split_amount",
                        field="amount",
                        message=("Both debit and credit values are populated"),
                        raw_value={
                            "debit": debit_value,
                            "credit": credit_value,
                        },
                    )
                )
                return None

            if not has_debit and not has_credit:
                issues.append(
                    ValidationIssue(
                        code="missing_amount",
                        field="amount",
                        message="A debit or credit amount is required",
                    )
                )
                return None

            if has_debit:
                parsed = self._parse_money_issue(
                    raw_value=debit_value,
                    field="debit_amount",
                    issues=issues,
                )
                amount = -abs(parsed) if parsed is not None else None
            else:
                parsed = self._parse_money_issue(
                    raw_value=credit_value,
                    field="credit_amount",
                    issues=issues,
                )
                amount = abs(parsed) if parsed is not None else None

        if amount == Decimal("0.00"):
            issues.append(
                ValidationIssue(
                    code="zero_amount",
                    field="amount",
                    message="A zero-value transaction requires manual review",
                    raw_value=amount,
                )
            )

        return amount

    def _parse_money_issue(
        self,
        *,
        raw_value: Any,
        field: str,
        issues: list[ValidationIssue],
    ) -> Decimal | None:
        if self._is_blank(raw_value):
            issues.append(
                ValidationIssue(
                    code="missing_amount",
                    field=field,
                    message="Transaction amount is required",
                    raw_value=raw_value,
                )
            )
            return None

        try:
            return self._parse_money(raw_value)
        except (InvalidOperation, TypeError, ValueError):
            issues.append(
                ValidationIssue(
                    code="invalid_amount",
                    field=field,
                    message="Amount is not a valid two-decimal monetary value",
                    raw_value=raw_value,
                )
            )
            return None

    @staticmethod
    def _parse_money(raw_value: Any) -> Decimal:
        if isinstance(raw_value, bool):
            raise TypeError("Boolean values are not monetary amounts")

        if isinstance(raw_value, Decimal):
            amount = raw_value
        elif isinstance(raw_value, int):
            amount = Decimal(raw_value)
        elif isinstance(raw_value, float):
            if not math.isfinite(raw_value):
                raise ValueError("Amount must be finite")
            amount = Decimal(str(raw_value))
        elif isinstance(raw_value, str):
            text = raw_value.strip()
            negative_parentheses = text.startswith("(") and text.endswith(")")

            if negative_parentheses:
                text = text[1:-1].strip()

            if text.upper().startswith("USD "):
                text = text[4:].strip()

            if text.startswith("$"):
                text = text[1:].strip()

            if not MONEY_PATTERN.fullmatch(text):
                raise ValueError("Unsupported money format")

            amount = Decimal(text.replace(",", ""))
            if negative_parentheses:
                amount = -abs(amount)
        else:
            raise TypeError("Unsupported amount type")

        if not amount.is_finite():
            raise ValueError("Amount must be finite")

        quantized = amount.quantize(CENT)
        if amount != quantized:
            raise ValueError("Amount contains more than two decimal places")

        return quantized

    def _date_field(
        self,
        *,
        raw_value: Any,
        field: str,
        date_format: str,
        required: bool,
        issues: list[ValidationIssue],
    ) -> date | None:
        if self._is_blank(raw_value):
            if required:
                issues.append(
                    ValidationIssue(
                        code="missing_date",
                        field=field,
                        message="Transaction date is required",
                        raw_value=raw_value,
                    )
                )
            return None

        try:
            if isinstance(raw_value, datetime):
                return raw_value.date()
            if isinstance(raw_value, date):
                return raw_value
            if isinstance(raw_value, str):
                return datetime.strptime(
                    raw_value.strip(),
                    date_format,
                ).date()
            raise TypeError("Unsupported date type")
        except (TypeError, ValueError):
            issues.append(
                ValidationIssue(
                    code="invalid_date",
                    field=field,
                    message=(f"Date does not match configured format {date_format}"),
                    raw_value=raw_value,
                )
            )
            return None

    def _currency_field(
        self,
        *,
        values: dict[str, Any],
        config: IngestionConfig,
        issues: list[ValidationIssue],
    ) -> str | None:
        mapping = config.column_mapping
        raw_value = values.get(mapping.currency) if mapping.currency else None

        if self._is_blank(raw_value):
            raw_value = config.default_currency

        if self._is_blank(raw_value):
            issues.append(
                ValidationIssue(
                    code="missing_currency",
                    field="currency",
                    message="Currency is required",
                    raw_value=raw_value,
                )
            )
            return None

        currency = str(raw_value).strip().upper()
        if len(currency) != 3 or not currency.isalpha():
            issues.append(
                ValidationIssue(
                    code="invalid_currency",
                    field="currency",
                    message="Currency must be a three-letter alphabetic code",
                    raw_value=raw_value,
                )
            )
            return None

        return currency

    def _bank_account_field(
        self,
        *,
        values: dict[str, Any],
        config: IngestionConfig,
        issues: list[ValidationIssue],
    ) -> str | None:
        mapping = config.column_mapping
        raw_value = values.get(mapping.bank_account) if mapping.bank_account else None

        if self._is_blank(raw_value):
            raw_value = config.default_bank_account

        bank_account = self._normalized_text(raw_value)
        if bank_account is None:
            issues.append(
                ValidationIssue(
                    code="missing_bank_account",
                    field="bank_account",
                    message="Bank account is required",
                    raw_value=raw_value,
                )
            )

        return bank_account

    @staticmethod
    def _optional_text(value: Any) -> str | None:
        if SourceFileNormalizer._is_blank(value):
            return None
        return str(value).strip()

    @staticmethod
    def _normalized_text(value: Any) -> str | None:
        if SourceFileNormalizer._is_blank(value):
            return None
        return " ".join(str(value).split())

    @staticmethod
    def _normalize_description(value: str | None) -> str | None:
        if value is None:
            return None

        normalized = unicodedata.normalize("NFKC", value)
        normalized = " ".join(normalized.split()).casefold()
        return normalized or None

    @staticmethod
    def _fingerprint(
        *,
        transaction_date: date | None,
        amount: Decimal | None,
        currency: str | None,
        bank_account: str | None,
        description: str | None,
    ) -> str | None:
        if (
            transaction_date is None
            or amount is None
            or currency is None
            or bank_account is None
            or description is None
        ):
            return None

        identity = {
            "transaction_date": transaction_date.isoformat(),
            "amount": format(amount, ".2f"),
            "currency": currency,
            "bank_account": bank_account.casefold(),
            "description": description,
        }
        encoded = json.dumps(
            identity,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    @staticmethod
    def _raw_hash(raw_values: dict[str, Any]) -> str:
        canonical = {
            key: SourceFileNormalizer._json_safe(value) for key, value in raw_values.items()
        }
        encoded = json.dumps(
            canonical,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    @staticmethod
    def _json_safe(value: Any) -> Any:
        if isinstance(value, datetime):
            return {"type": "datetime", "value": value.isoformat()}
        if isinstance(value, date):
            return {"type": "date", "value": value.isoformat()}
        if isinstance(value, Decimal):
            return {"type": "decimal", "value": str(value)}
        if isinstance(value, bytes):
            return {"type": "bytes", "value": value.hex()}
        if isinstance(value, float) and not math.isfinite(value):
            return {"type": "float", "value": repr(value)}
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return {
            "type": type(value).__name__,
            "value": str(value),
        }

    @staticmethod
    def _is_blank(value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())
